from gtts import gTTS
import json
from io import BytesIO
import streamlit as st
from openai import OpenAI
import numpy as np
import io, os
import fitz  # PyMuPDF
from docx import Document
import openpyxl

# openai_api_key = os.getenv("OPENAI_API_KEY")
# st.secrets["openai_api_key"] = os.getenv("OPENAI_API_KEY")
def get_doc_content(uploaded_file):
    # Determine file type and process accordingly
    if uploaded_file.name.endswith('.pdf'):
        return get_pdf_content(uploaded_file)
    elif uploaded_file.name.endswith('.txt'):
        return get_txt_content(uploaded_file)
    elif uploaded_file.name.endswith('.docx'):
        return get_docx_content(uploaded_file)
    elif uploaded_file.name.endswith('.xlsx'):
        return get_excel_content(uploaded_file)
    else:
        return "Unsupported file format"

def get_pdf_content(uploaded_file):
    content = ""
    with fitz.open(stream=uploaded_file.getvalue(), filetype="pdf") as doc:
        for page in doc:
            content += page.get_text()
    return content

def get_txt_content(uploaded_file):
    # For a text file, simply return its content
    return uploaded_file.getvalue().decode('utf-8')

def get_docx_content(uploaded_file):
    content = ""
    doc = Document(io.BytesIO(uploaded_file.getvalue()))
    for para in doc.paragraphs:
        content += para.text + "\n"
    return content

def get_excel_content(uploaded_file):
    content = ""
    workbook = openpyxl.load_workbook(filename=io.BytesIO(uploaded_file.getvalue()), data_only=True)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                content += str(cell) if cell is not None else "" + "\t"
            content += "\n"
    return content


def get_response(user_input:str, chat_history:str, doc_content:str):
    
    
    
    prompt = get_chat_prompt(user_input, chat_history, doc_content)
    
    return generate_text(prompt)



def get_chat_prompt(user_input:str, chat_history:str, doc_content:str):
    if len(doc_content) > 5:
        doc_state = """
        You have also been provided a document whose content you can use to answer the questions.
        You will find the answers to as many inteview questions in the document as possible and return whats mention in the document regarding which question.
        If the content is not relevant mention that.
        Finally add the question that takes forward the conversation.
        """
        doc_content = f"Document Content: {doc_content}"
    else:
        doc_state= ""
        doc_content = ""
    
    query = f"""
    You are having a conversation with a company rep. in the order to gather information about the company.
    You have been provided with the Chat history and users latest input.
    {doc_state}
    Your job is to drive the conversation forward in the order to complete the a questionaire.
    This questionaire is divided into two parts:
    1. Preliminary Data Gathering : This establishes basic information about the company
    2. Main Questions: Related to Context Section of ISO 9001:2015
    You will be asking a question from the questionaire based on the user's input.
    Associate relevant context to make your question more specific by including specifics mentioned by the user, like names, structural data, heirachies, and roles.
    You may also combine or simplify the questions to compensate for incomplete or too long information.
    If A previously asked question is partially answered, include the unanswered part in the next question. Explain how its important for the interview process.
    Add a small senetance explaining the progress (approx %) of the interview process.
    Here is the questionaire:
    ###
    Preliminary Data Gathering:
    1. What is the name of your organization? Can you provide the Organisation number and a point of contact?
    2. What does your organization do? (Main products, services, or activities)
    3. Who are your organization’s main customers?'
    4. Verify that the point of contact shall be verifying the information provided, if the user is the person or someone else.
    5  Can you briefly describe your organization’s history related to ISO 9001:2015 certification?
    6. Who leads your organization, and what is their role? Give Top Management names and Point of contact
    7. How many people work in organisation? Provide structure of responsibiliites or heirachy.
    8. What locations or regions does your organization operate in?
    
    Main Questions:

    9. Has the company evaluated the relevance of ISO certification for all its operational areas to determine if specific areas should not pursue certification?

    10. How is the company planning to introduce "systems thinking" and standardize work processes across its operations?

    11. What steps have been taken to ensure the board, management, and employees are all engaged and supportive of the ISO standards implementation?

    12. For each identified area of activity within the company, what are the specific market requirements, legal obligations, and customer expectations that must be met?

    13. What resources (both internal and external) are identified as necessary for meeting the requirements of each area of activity?

    14. Can you provide a detailed description of the processes in place for each area of activity, highlighting any sequential or parallel operations leading to delivery?

    15. What "soft resources" are prioritized for management to effectively implement and maintain ISO standards?

    16. How does the company plan to ensure the physical, social, and psychological environments are conducive to the operation of processes under ISO standards?

    17. What mechanisms are in place to ensure the monitoring and measurement resources are fit for purpose and capable of verifying product and service compliance?

    18. How is the organization ensuring that essential knowledge is maintained, made available, and updated according to changing needs and trends?

    19. What strategies are being employed to ascertain and enhance the competence of individuals performing work under the organization’s management?

    20. How does the company plan to manage both internal and external communications relevant to the quality management system, and what documented information will be included?
    ###
    Chat History:{chat_history}
    User's lastest input: {user_input}
    {doc_content}
    ***
    Respond as an interviewer who is direct and to the point, but also friendly and understanding.
    IF Previous question has been answered off-context, or partially or generally.
    Use markdown to highlight the Import tant parts of your output. 
    
    
    
    """
    return query
    


def generate_text(prompt:str):
    print("§§§§ $$$ CAUTION: A COSTING CALL#### IS BEING MADE $$$ §§§§")
    openai_api_key = os.getenv("OPENAI_API_KEY")
    client = OpenAI(
        
       api_key=openai_api_key
    )
    completion = client.chat.completions.create(
        model="gpt-4-turbo-preview",
        messages=[
            {
                "role": "system",  
                "content": "You are a vigilant information gathering audit assistant"
            },
            {
                "role": "user",
                "content": prompt,
            },
        ],
    )
    # print(type(completion.choices[0].message.content))
    return str(completion.choices[0].message.content)

def text_to_speech(text):
    tts = gTTS(text=text, lang='en')
    audio_buffer = BytesIO()
    tts.write_to_fp(audio_buffer)
    audio_buffer.seek(0)
    return audio_buffer


st.title("ISOGRAPH")
st.subheader("Helps gather preliminary Data for ISOEnsure")
# initalize session state


if "messages" not in st.session_state:
    st.session_state.messages = []
    
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = []



if len(st.session_state.messages) == 0:
    st.session_state.messages.append({"role": "assistant", "content": "Hi, I am going to help you OnBoard ISOEnsure. What is the name of your organization? Can you provide the Organisation number and a point of contact?"})
st.write("ISOGRAPH asks about 20 questions to gather preliminary data for ISOEnsure. The assistant may slightly modify its response based on your input. You can also upload a document to help answer the questions.")
"Consider ansering all of them to efficiently onboard ISOEnsure."
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

user_input = st.chat_input("Please provide the details requested above:")
if user_input:

    with st.chat_message("user"):
        st.markdown(user_input)

    st.session_state.messages.append({"role": "user", "content": user_input})

    with st.status(".", expanded=True):
        chat_history_str = " ".join([message["content"] for message in st.session_state.messages])
        # Assume get_response is a function you define elsewhere to generate the assistant's response
        # if there is a document uploaded get its content
        doc_content = ""
        if len(st.session_state.uploaded_files) > 0:
            for doc in st.session_state.uploaded_files:
                doc_content += get_doc_content(doc)
                doc_content += "\n"
        print(doc_content)
        
        response = get_response(user_input, chat_history_str, doc_content)  
        # Display assistant response in chat message container
        with st.chat_message("assistant"):
            st.markdown(response)
        # Add assistant response to chat history
        st.session_state.messages.append({"role": "assistant", "content": response})
        
        # play the audio as the response is generated
        audio_buffer = text_to_speech(response)
        st.audio(audio_buffer, format="audio/wav")

        
# Sidebar for chat history management
with st.sidebar:
    st.title("ISOGRAPH")
    st.subheader("OnBoarding Support for ISOEnsure")
    st.write("""
             Participate in the chat to gather information about the company.
             
             """)
    # Chat History Management
    "---"  

    if 'messages' in st.session_state:
        
        chat_history_json = json.dumps(st.session_state.messages, indent=2)
        # Convert it to string
        chat_history_str = " \n".join([message["content"] for message in st.session_state.messages])
        st.warning( "Remember to download the chat history, You will need it to start your ISOENSURE")
        st.download_button(
            label="Download CHAT HISTORY as OUPUT JSON",
            data=chat_history_json,
            file_name="chat_history.json",
            mime="application/json"
        )
        "---"
        if st.button("Clear Chat History"):
            st.session_state.messages = []
            st.experimental_rerun()
    
    # File upload and management
    "---"



    questions = """
    Preliminary Data Gathering:
    1. What is the name of your organization? Can you provide the Organisation number and a point of contact?
    2. What does your organization do? (Main products, services, or activities)
    3. Who are your organization’s main customers?'
    4. Verify that the point of contact shall be verifying the information provided, if the user is the person or someone else.
    5  Can you briefly describe your organization’s history related to ISO 9001:2015 certification?
    6. Who leads your organization, and what is their role? Give Top Management names and Point of contact
    7. How many people work in organisation? Provide structure of responsibiliites or heirachy.
    8. What locations or regions does your organization operate in?
    
    Main Questions:

    9. Has the company evaluated the relevance of ISO certification for all its operational areas to determine if specific areas should not pursue certification?

    10. How is the company planning to introduce "systems thinking" and standardize work processes across its operations?

    11. What steps have been taken to ensure the board, management, and employees are all engaged and supportive of the ISO standards implementation?

    12. For each identified area of activity within the company, what are the specific market requirements, legal obligations, and customer expectations that must be met?

    13. What resources (both internal and external) are identified as necessary for meeting the requirements of each area of activity?

    14. Can you provide a detailed description of the processes in place for each area of activity, highlighting any sequential or parallel operations leading to delivery?

    15. What "soft resources" are prioritized for management to effectively implement and maintain ISO standards?

    16. How does the company plan to ensure the physical, social, and psychological environments are conducive to the operation of processes under ISO standards?

    17. What mechanisms are in place to ensure the monitoring and measurement resources are fit for purpose and capable of verifying product and service compliance?

    18. How is the organization ensuring that essential knowledge is maintained, made available, and updated according to changing needs and trends?

    19. What strategies are being employed to ascertain and enhance the competence of individuals performing work under the organization’s management?

    20. How does the company plan to manage both internal and external communications relevant to the quality management system, and what documented information will be included?


"""
    st.download_button(
            label="Download Questions",
            data=questions,
            file_name="ISOENSURE_OnBoarding_Questions.txt",
            mime="text/plain"
        )
    "---"
    uploaded_file = st.file_uploader("Upload a Reference file:WORD, PDF, TXT supported ", key="file_uploader")
    if uploaded_file is not None:
        # Initialize the list if it doesn't exist
        if 'uploaded_files' not in st.session_state:
            st.session_state.uploaded_files = []
        
        # Add the uploaded file to the session state list
        st.session_state.uploaded_files.append(uploaded_file)